#!/usr/bin/env python3
"""Convert Douyin category sales JSON into an analyst-friendly Excel workbook.

Expected source shape:
{
  "Data": {
    "List": [
      {
        "DateCode": 20260427,
        "ListTimeStr": "2026/04/27",
        "Cates": [{ "DyCateId": "...", "DyCateName": "...", "SaleCount": 1, ... }]
      }
    ],
    "Total": [{ "DyCateId": "...", "DyCateName": "...", "SaleCount": 30, ... }]
  }
}
"""
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `python3 -m pip install openpyxl`."
    ) from exc


PERCENT_COLUMNS = {"销量占比", "GMV占比", "商品数占比", "SPU数占比"}
DATE_COLUMNS = {"日期", "开始日期", "结束日期"}
TEXT_COLUMNS = {"类目ID", "最高GMV类目ID", "最高销量类目ID", "Uid"}
INTEGER_COLUMNS = {
    "日期编码",
    "销量",
    "销售GMV",
    "商品数",
    "SPU数",
    "当日GMV排名",
    "当日销量排名",
    "GMV排名",
    "销量排名",
    "汇总销量",
    "汇总销售GMV",
    "类目数",
    "销量合计",
    "GMV合计",
    "商品数合计",
    "SPU数合计",
    "最高GMV",
    "最高销量",
}
DECIMAL_COLUMNS = {"件均GMV", "周期均销量", "周期均GMV"}
WIDTH_OVERRIDES = {
    "日期": 12,
    "开始日期": 12,
    "结束日期": 12,
    "周期": 24,
    "日期编码": 12,
    "类目ID": 12,
    "类目名称": 24,
    "销量区间": 14,
    "GMV区间": 14,
    "商品数文本": 12,
    "SPU数文本": 12,
    "日期范围": 24,
    "最高GMV类目": 24,
    "最高销量类目": 24,
    "检查项": 18,
    "值": 42,
    "说明": 70,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert Data.List/Data.Total category sales JSON to a formatted .xlsx workbook."
    )
    parser.add_argument("input", type=Path, help="Source JSON path")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output .xlsx path. Defaults to outputs/<input>_category_sales_<start>_<end>.xlsx",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit non-zero if Data.Total does not reconcile with daily detail totals.",
    )
    return parser.parse_args()


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(value, fmt)
            return parsed.date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date format: {value!r}")


def parse_period(value: str | None, fallback_label: str | None = None) -> tuple[date | None, date | None, str]:
    """Parse a single date or a date range.

    Supported examples:
    - 2026/04/27
    - 2026-04-27
    - 2026/01/01-2026/01/31
    - 2026-01-01~2026-01-31
    """
    label = value or fallback_label or ""
    if not value:
        return None, None, label

    normalized = value.replace("~", "-")
    # Split only between two date tokens, not inside yyyy-mm-dd.
    if "/" in normalized and "-" in normalized:
        start_text, end_text = normalized.split("-", 1)
        return parse_date(start_text.strip()), parse_date(end_text.strip()), label
    if normalized.count("-") == 1 and "/" not in normalized:
        start_text, end_text = normalized.split("-", 1)
        return parse_date(start_text.strip()), parse_date(end_text.strip()), label

    single = parse_date(value.strip())
    return single, single, label


def safe_number(value: Any) -> int | float:
    return value if isinstance(value, (int, float)) else 0


def divide(numerator: int | float, denominator: int | float) -> float | None:
    return round(numerator / denominator, 4) if denominator else None


def load_payload(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as source:
        payload = json.load(source)
    if not isinstance(payload, dict):
        raise ValueError("JSON root must be an object.")
    data = payload.get("Data")
    if not isinstance(data, dict) or not isinstance(data.get("List"), list):
        raise ValueError("JSON must contain Data.List as a list.")
    return payload


def compute_totals_from_detail(days: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals: dict[tuple[str, str], dict[str, Any]] = {}
    for day in days:
        for cate in day.get("Cates") or []:
            if not isinstance(cate, dict):
                continue
            key = (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or ""))
            row = totals.setdefault(
                key,
                {
                    "Uid": cate.get("Uid"),
                    "DyCateId": key[0],
                    "DyCateName": key[1],
                    "SaleCount": 0,
                    "SaleGmv": 0,
                    "GoodCount": 0,
                    "SpuCount": 0,
                },
            )
            for metric in ("SaleCount", "SaleGmv", "GoodCount", "SpuCount"):
                row[metric] += safe_number(cate.get(metric))
    return list(totals.values())


def build_rows(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], bool]:
    data = payload["Data"]
    days = [item for item in data.get("List", []) if isinstance(item, dict)]
    source_totals = data.get("Total") or compute_totals_from_detail(days)
    totals = [item for item in source_totals if isinstance(item, dict)]
    periods = [parse_period(day.get("ListTimeStr"), day.get("PicTimeStr")) for day in days]
    start_dates = [start for start, _, _ in periods if start is not None]
    end_dates = [end for _, end, _ in periods if end is not None]
    date_range = (
        f"{min(start_dates).isoformat()}~{max(end_dates).isoformat()}"
        if start_dates and end_dates
        else ""
    )

    detail_rows: list[dict[str, Any]] = []
    for day in days:
        start_date, end_date, period_label = parse_period(day.get("ListTimeStr"), day.get("PicTimeStr"))
        cates = [item for item in day.get("Cates") or [] if isinstance(item, dict)]
        gmv_rank = {
            (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or "")): index + 1
            for index, cate in enumerate(sorted(cates, key=lambda item: safe_number(item.get("SaleGmv")), reverse=True))
        }
        count_rank = {
            (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or "")): index + 1
            for index, cate in enumerate(sorted(cates, key=lambda item: safe_number(item.get("SaleCount")), reverse=True))
        }
        for cate in cates:
            key = (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or ""))
            sale_count = safe_number(cate.get("SaleCount"))
            sale_gmv = safe_number(cate.get("SaleGmv"))
            detail_rows.append(
                {
                    "周期": period_label,
                    "开始日期": start_date,
                    "结束日期": end_date,
                    "日期": start_date,
                    "日期编码": day.get("DateCode"),
                    "类目ID": key[0],
                    "类目名称": key[1],
                    "销量": cate.get("SaleCount"),
                    "销量区间": cate.get("SaleCountStr"),
                    "销量占比": cate.get("SaleCountRatio"),
                    "销量占比文本": cate.get("SaleCountRatioStr"),
                    "销售GMV": cate.get("SaleGmv"),
                    "GMV区间": cate.get("SaleGmvStr"),
                    "GMV占比": cate.get("SaleGmvRatio"),
                    "GMV占比文本": cate.get("SaleGmvRatioStr"),
                    "商品数": cate.get("GoodCount"),
                    "商品数文本": cate.get("GoodCountStr"),
                    "商品数占比": cate.get("GoodCountRatio"),
                    "商品数占比文本": cate.get("GoodCountRatioStr"),
                    "SPU数": cate.get("SpuCount"),
                    "SPU数文本": cate.get("SpuCountStr"),
                    "SPU数占比": cate.get("SpuCountRatio"),
                    "SPU数占比文本": cate.get("SpuCountRatioStr"),
                    "件均GMV": divide(sale_gmv, sale_count),
                    "当日GMV排名": gmv_rank[key],
                    "当日销量排名": count_rank[key],
                    "Uid": cate.get("Uid"),
                }
            )

    summary_source = sorted(totals, key=lambda item: safe_number(item.get("SaleGmv")), reverse=True)
    count_rank = {
        (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or "")): index + 1
        for index, cate in enumerate(sorted(totals, key=lambda item: safe_number(item.get("SaleCount")), reverse=True))
    }
    summary_rows: list[dict[str, Any]] = []
    day_count = len(days)
    for index, cate in enumerate(summary_source, start=1):
        key = (str(cate.get("DyCateId") or ""), str(cate.get("DyCateName") or ""))
        sale_count = safe_number(cate.get("SaleCount"))
        sale_gmv = safe_number(cate.get("SaleGmv"))
        summary_rows.append(
            {
                "GMV排名": index,
                "销量排名": count_rank[key],
                "类目ID": key[0],
                "类目名称": key[1],
                "汇总销量": cate.get("SaleCount"),
                "销量区间": cate.get("SaleCountStr"),
                "销量占比": cate.get("SaleCountRatio"),
                "销量占比文本": cate.get("SaleCountRatioStr"),
                "汇总销售GMV": cate.get("SaleGmv"),
                "GMV区间": cate.get("SaleGmvStr"),
                "GMV占比": cate.get("SaleGmvRatio"),
                "GMV占比文本": cate.get("SaleGmvRatioStr"),
                "商品数": cate.get("GoodCount"),
                "商品数区间": cate.get("GoodCountStr"),
                "商品数占比": cate.get("GoodCountRatio"),
                "商品数占比文本": cate.get("GoodCountRatioStr"),
                "SPU数": cate.get("SpuCount"),
                "SPU数区间": cate.get("SpuCountStr"),
                "SPU数占比": cate.get("SpuCountRatio"),
                "SPU数占比文本": cate.get("SpuCountRatioStr"),
                "周期均销量": divide(sale_count, day_count),
                "周期均GMV": divide(sale_gmv, day_count),
                "件均GMV": divide(sale_gmv, sale_count),
                "日期范围": date_range,
                "Uid": cate.get("Uid"),
            }
        )

    daily_total_rows: list[dict[str, Any]] = []
    for day in days:
        cates = [item for item in day.get("Cates") or [] if isinstance(item, dict)]
        sale_count = sum(safe_number(cate.get("SaleCount")) for cate in cates)
        sale_gmv = sum(safe_number(cate.get("SaleGmv")) for cate in cates)
        top_gmv = max(cates, key=lambda item: safe_number(item.get("SaleGmv")), default={})
        top_count = max(cates, key=lambda item: safe_number(item.get("SaleCount")), default={})
        daily_total_rows.append(
            {
                "周期": parse_period(day.get("ListTimeStr"), day.get("PicTimeStr"))[2],
                "开始日期": parse_period(day.get("ListTimeStr"), day.get("PicTimeStr"))[0],
                "结束日期": parse_period(day.get("ListTimeStr"), day.get("PicTimeStr"))[1],
                "日期": parse_period(day.get("ListTimeStr"), day.get("PicTimeStr"))[0],
                "日期编码": day.get("DateCode"),
                "类目数": len(cates),
                "销量合计": sale_count,
                "GMV合计": sale_gmv,
                "商品数合计": sum(safe_number(cate.get("GoodCount")) for cate in cates),
                "SPU数合计": sum(safe_number(cate.get("SpuCount")) for cate in cates),
                "件均GMV": divide(sale_gmv, sale_count),
                "最高GMV类目ID": str(top_gmv.get("DyCateId") or ""),
                "最高GMV类目": top_gmv.get("DyCateName"),
                "最高GMV": top_gmv.get("SaleGmv"),
                "最高销量类目ID": str(top_count.get("DyCateId") or ""),
                "最高销量类目": top_count.get("DyCateName"),
                "最高销量": top_count.get("SaleCount"),
            }
        )

    detail_sale_count = sum(safe_number(row.get("销量")) for row in detail_rows)
    detail_sale_gmv = sum(safe_number(row.get("销售GMV")) for row in detail_rows)
    total_sale_count = sum(safe_number(row.get("SaleCount")) for row in totals)
    total_sale_gmv = sum(safe_number(row.get("SaleGmv")) for row in totals)
    reconciled = detail_sale_count == total_sale_count and detail_sale_gmv == total_sale_gmv

    name_to_ids: dict[str, set[str]] = defaultdict(set)
    for row in totals:
        name_to_ids[str(row.get("DyCateName") or "")].add(str(row.get("DyCateId") or ""))
    duplicate_names = [
        f"{name}({'/'.join(sorted(ids))})"
        for name, ids in sorted(name_to_ids.items())
        if name and len(ids) > 1
    ]
    per_day_counts = sorted({len(day.get("Cates") or []) for day in days})
    validation_rows = [
        {"检查项": "接口状态", "值": str(payload.get("Status")), "说明": f"Code={payload.get('Code')}; Msg={payload.get('Msg')!r}"},
        {"检查项": "日期范围", "值": date_range, "说明": "按 Data.List 解析"},
        {"检查项": "周期数", "值": len(days), "说明": "Data.List 记录数"},
        {"检查项": "每周期类目数", "值": ",".join(map(str, per_day_counts)), "说明": "每个周期的 Cates 数量"},
        {"检查项": "周期明细行数", "值": len(detail_rows), "说明": "周期数 × 每周期类目数"},
        {"检查项": "汇总行数", "值": len(totals), "说明": "Data.Total 记录数；缺失时由周期明细生成"},
        {"检查项": "销量校验", "值": "一致" if detail_sale_count == total_sale_count else "不一致", "说明": f"周期明细销量合计={detail_sale_count}; 汇总销量合计={total_sale_count}"},
        {"检查项": "GMV校验", "值": "一致" if detail_sale_gmv == total_sale_gmv else "不一致", "说明": f"周期明细GMV合计={detail_sale_gmv}; 汇总GMV合计={total_sale_gmv}"},
        {"检查项": "类目唯一键", "值": "类目ID+类目名称", "说明": "避免同名不同ID类目被误合并"},
        {"检查项": "同名不同ID类目", "值": "; ".join(duplicate_names) if duplicate_names else "无", "说明": "保留为独立类目"},
        {"检查项": "数值处理", "值": "未换算、未四舍五入原始销量/GMV", "说明": "SaleCount 和 SaleGmv 使用 JSON 原始数值；区间字段另列保留"},
        {"检查项": "生成时间", "值": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "说明": "本地生成 Excel 时间"},
    ]
    return detail_rows, summary_rows, daily_total_rows, validation_rows, reconciled


def add_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title=name)
    if not rows:
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])

    max_row = worksheet.max_row
    max_col = worksheet.max_column
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = headers[cell.column - 1]
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if header in DATE_COLUMNS and cell.value:
                cell.number_format = "yyyy-mm-dd"
            elif header in PERCENT_COLUMNS and cell.value is not None:
                cell.number_format = "0.00%"
            elif header in INTEGER_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            elif header in DECIMAL_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            elif header in TEXT_COLUMNS and cell.value is not None:
                cell.number_format = "@"

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        max_len = len(str(header))
        sample_cells = worksheet[letter][1 : min(max_row, 80)]
        for cell in sample_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(WIDTH_OVERRIDES.get(header, max_len + 2), 10), 70)


def default_output_path(input_path: Path, detail_rows: list[dict[str, Any]]) -> Path:
    start_dates = [row["开始日期"] for row in detail_rows if isinstance(row.get("开始日期"), date)]
    end_dates = [row["结束日期"] for row in detail_rows if isinstance(row.get("结束日期"), date)]
    if start_dates and end_dates:
        suffix = f"{min(start_dates).strftime('%Y%m%d')}_{max(end_dates).strftime('%Y%m%d')}"
    else:
        suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("outputs") / f"{input_path.stem}_category_sales_{suffix}.xlsx"


def write_workbook(output_path: Path, detail_rows: list[dict[str, Any]], summary_rows: list[dict[str, Any]], daily_rows: list[dict[str, Any]], validation_rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "周期销售明细", detail_rows)
    add_sheet(workbook, "类目汇总", summary_rows)
    add_sheet(workbook, "周期总计", daily_rows)
    add_sheet(workbook, "数据校验", validation_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    payload = load_payload(args.input)
    detail_rows, summary_rows, daily_rows, validation_rows, reconciled = build_rows(payload)
    output_path = args.output or default_output_path(args.input, detail_rows)
    write_workbook(output_path, detail_rows, summary_rows, daily_rows, validation_rows)

    print(f"Excel written: {output_path.resolve()}")
    print(f"Rows: detail={len(detail_rows)}, summary={len(summary_rows)}, daily={len(daily_rows)}")
    print(f"Reconciled: {reconciled}")
    if args.strict and not reconciled:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
